from django.shortcuts import render, redirect
from unfold.views import UnfoldModelAdminViewMixin
from django.views.generic import TemplateView
from .models import Metakinhsh, OfficeSchedule
from django.urls import reverse
from django.http import JsonResponse, HttpResponse
from docxtpl import DocxTemplate
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from django.contrib import messages
from django.contrib.admin import site
from django.db.models.functions import ExtractMonth
from io import BytesIO
from datetime import datetime
from app.utils import is_member
from django.conf import settings


# Create your views here.
class MetakinhshCustomView(UnfoldModelAdminViewMixin, TemplateView):
    title = "Ημερολόγιο μετακινήσεων"
    template_name = "admin/index_metak.html"
    permission_required = "metakinhseis.view_metakinhsh"

    def get_context_data(self, **kwargs):
        is_elevated = self.request.user.is_superuser or is_member(self.request.user, 'Επόπτες')
        metak_cnt = Metakinhsh.objects.all().count() if  is_elevated \
            else Metakinhsh.objects.filter(consultant=self.request.user).count()
        unapproved_cnt = Metakinhsh.objects.filter(egkrish=False).count() if is_elevated \
            else Metakinhsh.objects.filter(consultant=self.request.user, egkrish=False).count()
        incomplete_cnt = Metakinhsh.objects.filter(pragmat=False).count() if is_elevated \
            else Metakinhsh.objects.filter(consultant=self.request.user, pragmat=False).count()
        complete_cnt = Metakinhsh.objects.filter(pragmat=True).count() if is_elevated \
            else Metakinhsh.objects.filter(consultant=self.request.user, pragmat=True).count()
        
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        # Add in a QuerySet of all the Metakinhseis
        context["kpi"] = [
            {
                "title": "Μετακινήσεις",
                "metric": metak_cnt,
            },
            {
                "title": "Χωρίς έγκριση",
                "metric": unapproved_cnt,
            },
            {
                "title": "Μη ολοκληρωμένες",
                "metric": incomplete_cnt,
            },
            {
                "title": "Ολοκληρωμένες",
                "metric": complete_cnt,
            },
        ]
        context['SHOW_OFFICE_DAYS'] = getattr(settings, 'SHOW_OFFICE_DAYS', False)  # Default to False if not set
        
        return context


# Get evaluation steps in JSON for the FullCalendar view
def metakinhsh_json(request):
    all_events = []
    events = []
    is_consultant = request.user.groups.filter(name='Σύμβουλοι').exists()
    metakinhseis = Metakinhsh.objects.all() if not is_consultant else Metakinhsh.objects.filter(consultant=request.user)

    for metak in metakinhseis:
        if is_consultant:
            description = f'{metak.metak_to}. Λόγος: {metak.aitiologia}'
        else:
            description = f'{metak.consultant.last_name}. Λόγος: {metak.aitiologia}'
        step_url = reverse('admin:metakinhseis_metakinhsh_change', args=[metak.id])

        events.append({
            'title': description,
            'start': metak.date_from.isoformat(),
            'end': metak.date_to.isoformat(),  # Use a different field if it's a range
            'allDay': True,  # Adjust if steps span time,
            'description': description,
            'url': step_url,
            'egkrish': metak.egkrish,
            'pragmat': metak.pragmat
        })
    all_events = events

    # If SHOW_OFFICE_DAYS are enabled
    if settings.SHOW_OFFICE_DAYS:
        office_days_events = []
        greek_months = [
                    "Ιανουάριος", "Φεβρουάριος", "Μάρτιος",
                    "Απρίλιος", "Μάιος", "Ιούνιος",
                    "Ιούλιος", "Αύγουστος", "Σεπτέμβριος",
                    "Οκτώβριος", "Νοέμβριος", "Δεκέμβριος"
                ]
        start = request.GET['start']
        end = request.GET['end']

        startdt = datetime.strptime(start[:10], '%Y-%m-%d')
        enddt = datetime.strptime(end[:10], '%Y-%m-%d')
        
        # Find the middle date of the range
        middle_date = startdt + (enddt - startdt) / 2
        cur_month = middle_date.month
        cur_year = middle_date.year
        current_month = f'{greek_months[cur_month-1]} {cur_year}'
        
        # Return consultant's records or all
        if is_consultant:
            records = OfficeSchedule.objects.filter(consultant=request.user, month=current_month)
        else:
            records = OfficeSchedule.objects.filter(month=current_month)

        # Set office days to events
        for rec in records:
            for day in rec.days_in_office:
                title = 'Ημέρα γραφείου' if is_consultant else f'{rec.consultant.last_name} {rec.consultant.first_name[:1]}.'
                office_days_events.append({
                    'title': title,
                    'start': day,
                    'allDay': True,
                    'description': 'Ημέρα γραφείου',
                    'officeDay': True
                })
        all_events = events + office_days_events
    return JsonResponse(all_events, safe=False)


def apofasi_metakinhshs_preview(request):
    metakinhsh_ids_str = request.GET.get('metakinhsh_ids', '')
    is_oikon = request.GET.get('oikon','')
    metakinhsh_ids = [int(id.strip()) for id in metakinhsh_ids_str.split(',') if id.strip().isdigit()]

    # If the form is submitted
    if request.method == "POST":
        metakinhsh_ids_str = request.POST.get('metakinhsh_ids', '')
        metakinhsh_ids = [int(id.strip()) for id in metakinhsh_ids_str.split(',') if id.strip().isdigit()]

        if not metakinhsh_ids:
            messages.error(request, "Σφάλμα: Δε δόθηκαν Α/Α μετακινήσεων.")
            return redirect(reverse('admin:metakinhseis_metakinhsh_changelist'))  # Redirect back to user list

        try:
            metakinhseis = Metakinhsh.objects.filter(id__in=metakinhsh_ids)
        except Metakinhsh.DoesNotExist:
            messages.error(request, f"Η μετακίνηση δε βρέθηκε.")
            return redirect(reverse('admin:metakinhseis_metakinhsh_changelist'))  # Redirect back to user list
        
        # Start constructing document & its data
        docx_template = 'metakinhseis/templates/word/tmpl_apof_metak_oikon.docx' if is_oikon else 'metakinhseis/templates/word/tmpl_apof_metak.docx'
        doc = DocxTemplate(docx_template)
        results = list()
        consultants = list()
        aa = 1
        if is_oikon:
            total_days = 0
            total_km = 0
            total_amount_ektos = 0.0
            total_amount_km = 0.0

        for met in metakinhseis:
            row = dict()
            row['aa'] = aa
            row['date_from'] = met.date_from.strftime("%d-%m-%Y")
            row['aitiologia'] = met.aitiologia
            row['transport'] = 'Ι.Χ.'
            if is_oikon:
                row['km'] = met.km
                row['metak_to'] = met.metak_to
                total_days += 1
                total_km += met.km * 2
                total_amount_ektos += 10.0
                total_amount_km += (met.km * 2) * 0.2

            consultants.append(met.consultant.id)
            results.append(row)
            aa += 1

        # check if at least one metakinhsh is for another consultant and throw error
        if len(set(consultants)) > 1:
            messages.error(request, "Σφάλμα: Οι επιλεγμένες μετακινήσεις δεν αφορούν μόνο σε έναν σύμβουλο.")
            return redirect(reverse('admin:metakinhseis_metakinhsh_changelist'))  # Redirect back to user list
        try:    
            context = {
                'surname': met.consultant.last_name,
                'name': met.consultant.first_name,
                'klados': met.consultant.consultant.klados,
                'enothta': met.consultant.consultant.enothta,
                'table_data': results
            }
        except:
            messages.error(request, "Σφάλμα: Δεν έχουν εισαχθεί τα επιπλέον στοιχεία (κλάδος, ενότητα) για την/ην σύμβουλο.")
            return redirect(reverse('admin:metakinhseis_metakinhsh_changelist'))  # Redirect back to user list
        if is_oikon:
            context.update({
                'total_amount_ektos': round(total_amount_ektos, 2),
                'total_amount_km': round(total_amount_km, 2),
                'total_km': total_km,
                'total_days': total_days
            })
        # left for debugging
        # print(context)
        # return redirect(reverse('admin:metakinhseis_metakinhsh_changelist'))  # Redirect to the user list page after assignment
        doc.render(context)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
        response['Content-Disposition'] = 'attachment; filename=apof_metak.docx'
        doc.save(response)
        messages.success(request, 'Επιτυχής δημιουργία απόφασης μετακίνησης!')
        return response

    # If it's a GET request, render the form
    metakinhseis = Metakinhsh.objects.filter(id__in=metakinhsh_ids)

    # Perform checks
    # Check if approved
    if metakinhseis.filter(egkrish=False).exists():
        messages.error(request, f"Σφάλμα: Μία τουλάχιστον μετακίνηση δεν έχει εγκριθεί.")
        return redirect(reverse('admin:metakinhseis_metakinhsh_changelist'))
    # Check if done
    if metakinhseis.filter(pragmat=False).exists():
        messages.error(request, f"Σφάλμα: Μία τουλάχιστον μετακίνηση δεν έχει πραγματοποιηθεί.")
        return redirect(reverse('admin:metakinhseis_metakinhsh_changelist'))
    # Check if >50 km and not is_evaluation
    metakinhseis_unpaid = metakinhseis.filter(km__lte=50, is_evaluation__exact=False)
    metakinhseis_unpaid_ids = metakinhseis_unpaid.values_list('id', flat=True)
    if metakinhseis_unpaid_ids:
        messages.error(request, f"Σφάλμα: H/Οι παρακάτω μετακίνηση(-εις) {list(metakinhseis_unpaid_ids)} είναι <50 χλμ.")
        return redirect(reverse('admin:metakinhseis_metakinhsh_changelist'))
    
    # Check if <5 km and is_evaluation
    metakinhseis_unpaid = metakinhseis.filter(km__lte=5, is_evaluation__exact=True)
    metakinhseis_unpaid_ids = metakinhseis_unpaid.values_list('id', flat=True)
    if metakinhseis_unpaid_ids:
        messages.error(request, f"Σφάλμα: H/Οι παρακάτω μετακίνηση(-εις) αξιολόγησης {list(metakinhseis_unpaid_ids)} είναι <5 χλμ.")
        return redirect(reverse('admin:metakinhseis_metakinhsh_changelist'))
    
    context = {'metakinhseis': metakinhseis, 'metakinhsh_ids': metakinhsh_ids}
    context.update(site.each_context(request))

    return render(request, 'admin/preview_apofasi_metakinhshs.html', context)


def katastash_plhrwmhs(request):
    metakinhsh_ids_str = request.GET.get('metakinhsh_ids', '')
    metakinhsh_ids = [int(id.strip()) for id in metakinhsh_ids_str.split(',') if id.strip().isdigit()]

    # If the form is submitted
    if request.method == "POST":
        metakinhsh_ids_str = request.POST.get('metakinhsh_ids', '')
        metakinhsh_ids = [int(id.strip()) for id in metakinhsh_ids_str.split(',') if id.strip().isdigit()]

        if not metakinhsh_ids:
            messages.error(request, "Σφάλμα: Δε δόθηκαν Α/Α μετακινήσεων.")
            return redirect(reverse('admin:metakinhseis_metakinhsh_changelist'))  # Redirect back to user list

        try:
            metakinhseis = Metakinhsh.objects.filter(id__in=metakinhsh_ids)
        except Metakinhsh.DoesNotExist:
            messages.error(request, f"Η μετακίνηση δε βρέθηκε.")
            return redirect(reverse('admin:metakinhseis_metakinhsh_changelist'))  # Redirect back to user list
        
        # Start constructing document & its data

        # Number of rows to insert and the row index after which to insert
        n = len(metakinhseis)  # Number of rows to insert
        insert_after = 12  # Insert rows after this row (1-based index)

        # Load the Excel template
        template_path = 'metakinhseis/templates/excel/tmpl_katastash_plhrwmhs.xlsx' 
        # output_path = 'katastash_plhrwmhs.xlsx'

        # Open the workbook
        workbook = load_workbook(template_path)
        sheet = workbook.active  # Assuming you're working on the first sheet

        results = list()
        consultants = list()
        aa = 1
        
        total_days = 0
        total_km = 0
        total_amount_ektos = 0.0
        total_amount_km = 0.0
        months = list()

        for met in metakinhseis:
            row = dict()
            row['aa'] = aa
            row['date_from'] = met.date_from.strftime("%d-%m-%Y")
            row['aitiologia'] = met.aitiologia
            row['transport'] = 'Ι.Χ.'
            row['km'] = met.km
            row['metak_from'] = met.metak_from
            row['metak_to'] = met.metak_to
            total_days += 1
            total_km += met.km * 2
            total_amount_ektos += 10.0
            total_amount_km += (met.km * 2) * 0.2

            months.append(met.date_from.strftime('%m/%Y'))
            consultants.append(met.consultant.id)
            results.append(row)
            aa += 1

        # check if at least one metakinhsh is for another consultant and throw error
        if len(set(consultants)) > 1:
            messages.error(request, "Σφάλμα: Οι επιλεγμένες μετακινήσεις δεν αφορούν μόνο σε έναν σύμβουλο.")
            return redirect(reverse('admin:metakinhseis_metakinhsh_changelist'))  # Redirect back to user list
        
        total_amount = total_amount_ektos + total_amount_km
        total_mtpy = total_days * 0.2
        general_total = total_amount - total_mtpy
        # convert months to set to de-duplicate
        month_set = set(months)
        
        # Construct data to be replaced in xlsx
        try:
            context = {
                'surname': met.consultant.last_name,
                'name': met.consultant.first_name,
                'klados': met.consultant.consultant.klados,
                'enothta': met.consultant.consultant.enothta,
                'am': met.consultant.consultant.am,
                'iban': met.consultant.consultant.iban,
                'afm': met.consultant.username,
                'cur_date': datetime.now().strftime("%d-%m-%Y"),
                'months': ', '.join(month_set),
                'total_amount_ektos': round(total_amount_ektos, 2),
                'total_amount_km': round(total_amount_km, 2),
                'total_km': total_km,
                'total_days': total_days,
                'total_amount': total_amount,
                'total_mtpy': total_mtpy,
                'general_total': general_total
            }
        except Exception as e:
            print(e)
            messages.error(request, "Σφάλμα: Δεν έχουν εισαχθεί τα επιπλέον στοιχεία (κλάδος, ενότητα) για την/ην σύμβουλο.")
            return redirect(reverse('admin:metakinhseis_metakinhsh_changelist'))  # Redirect back to user list
        
        # Insert n rows after a specific row
        sheet.insert_rows(insert_after + 1, n)
        # Fill the inserted rows with data
        i = insert_after + 1
        for res in results:
            sheet.cell(row=i, column=1).value = res['date_from']
            sheet.cell(row=i, column=2).value = res['metak_from'] + ' - ' + res['metak_to']
            sheet.cell(row=i, column=3).value = res['aitiologia']
            km = res['km'] * 2
            sheet.cell(row=i, column=4).value = km
            sheet.cell(row=i, column=5).value = 'I.X.'
            total = km * 0.2
            sheet.cell(row=i, column=6).value = total
            sheet.cell(row=i, column=7).value = '0,00'
            sheet.cell(row=i, column=8).value = 0
            sheet.cell(row=i, column=9).value = '0,00'
            sheet.cell(row=i, column=10).value = 1
            sheet.cell(row=i, column=11).value = 10
            sheet.cell(row=i, column=12).value = total
            sheet.cell(row=i, column=13).value = '0,00'
            sheet.cell(row=i, column=14).value = 10
            sheet.cell(row=i, column=15).value = total + 10
            sheet.cell(row=i, column=16).value = '0.2'
            sheet.cell(row=i, column=17).value = total + 10 - 0.2
            i += 1

        # Replace placeholders in the sheet
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):  # Check if the cell contains a string
                    for key, val in context.items():
                        placeholder = f'{{{{{key}}}}}'  # e.g., "{{name}}"
                        if placeholder in cell.value:
                            cell.value = cell.value.replace(placeholder, str(val))

        # Define border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        # Apply border to all cells in the sheet
        for row in sheet.iter_rows(min_row=insert_after+1, max_row=insert_after+1+n):
            for cell in row:
                if cell.value is not None:  # Optional: Only apply border to cells with data
                    cell.border = thin_border

        # Save the updated workbook
        # Save to an in-memory file
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        # Create response to send the file to the user
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=katastash_plhrwmhs.xlsx'
        messages.success(request, 'Επιτυχής δημιουργία κατάστασης πληρωμής!')
        return response


    # If it's a GET request, render the form
    metakinhseis = Metakinhsh.objects.filter(id__in=metakinhsh_ids)

    # Perform checks
    # Check if approved
    if metakinhseis.filter(egkrish=False).exists():
        messages.error(request, f"Σφάλμα: Μία τουλάχιστον μετακίνηση δεν έχει εγκριθεί.")
        return redirect(reverse('admin:metakinhseis_metakinhsh_changelist'))
    # Check if done
    if metakinhseis.filter(pragmat=False).exists():
        messages.error(request, f"Σφάλμα: Μία τουλάχιστον μετακίνηση δεν έχει πραγματοποιηθεί.")
        return redirect(reverse('admin:metakinhseis_metakinhsh_changelist'))
    # Check if >50 km and not is_evaluation
    metakinhseis_unpaid = metakinhseis.filter(km__lte=50, is_evaluation__exact=False)
    metakinhseis_unpaid_ids = metakinhseis_unpaid.values_list('id', flat=True)
    if metakinhseis_unpaid_ids:
        messages.error(request, f"Σφάλμα: H/Οι παρακάτω μετακίνηση(-εις) {list(metakinhseis_unpaid_ids)} είναι <50 χλμ.")
        return redirect(reverse('admin:metakinhseis_metakinhsh_changelist'))
    
    # Check if <5 km and is_evaluation
    metakinhseis_unpaid = metakinhseis.filter(km__lte=5, is_evaluation__exact=True)
    metakinhseis_unpaid_ids = metakinhseis_unpaid.values_list('id', flat=True)
    if metakinhseis_unpaid_ids:
        messages.error(request, f"Σφάλμα: H/Οι παρακάτω μετακίνηση(-εις) αξιολόγησης {list(metakinhseis_unpaid_ids)} είναι <5 χλμ.")
        return redirect(reverse('admin:metakinhseis_metakinhsh_changelist'))
    
    # check if all metakinhseis were in the same month
    months = set(metakinhseis.annotate(metak_date=ExtractMonth('date_from')).values_list('metak_date', flat=True))
    if len(months) > 1:
        messages.error(request, "Σφάλμα: Οι μετακίνησεις που επιλέχθηκαν δεν πραγματοποιήθηκαν τον ίδιο μήνα")
        return redirect(reverse('admin:metakinhseis_metakinhsh_changelist'))
    
    context = {'metakinhseis': metakinhseis, 'metakinhsh_ids': metakinhsh_ids}
    context.update(site.each_context(request))

    return render(request, 'admin/preview_apofasi_metakinhshs.html', context)
